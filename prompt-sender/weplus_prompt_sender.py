"""
WePlus Prompt Sender v10
Local Windows desktop helper for supervised prompt sending.

v8 changes:
- One project queue file: 1 shot = 1 row with still + motion prompts together.
- v9: More robust Excel import: finds header row, supports Project Queue sheet, clearer openpyxl/.xls errors.
- Phase selector: Still or Motion.
- Global refs: Add Product / Face / Outfit / Style, Clear Refs.
- v10: prevents duplicate refs by using either global refs OR per-row refs, never both.
- v11: removed Apply All/Selected from the main flow. Global refs are session-level and auto-apply to still jobs.
- Removed Apply Pending to reduce confusion.
- Multi-click sequence: last click step is the final paste target.
- No cached sent status across Replace CSV imports; Replace resets statuses to pending.
- Autosave goes to *.autosave.csv, never overwrites source queue automatically.
"""
from __future__ import annotations

import csv
import json
import os
import queue
import struct
import threading
import time
import tkinter as tk
from dataclasses import dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Iterable

APP_TITLE = "WePlus Prompt Sender"
APP_DIR = Path(__file__).resolve().parent
PROFILE_PATH = APP_DIR / "sender_profiles.json"
LOG_PATH = APP_DIR / "sender-debug.log"
CF_HDROP = 15
CF_UNICODETEXT = 13
GMEM_MOVEABLE = 0x0002
STOP_MESSAGE = "Stopped by user"

PROJECT_FIELDNAMES = [
    "enabled", "shot_order", "shot_id", "shot_type", "category_th", "scene_preset", "scene_custom",
    "style_preset", "aspect_ratio", "profile_still", "profile_motion", "still_prompt", "motion_prompt",
    "attachments", "product_ref_names", "face_ref_names", "outfit_ref_names", "style_ref_names",
    "still_delay_sec", "motion_delay_sec", "still_output_name", "first_frame_path", "motion_output_name",
    "still_status", "motion_status", "notes",
]

try:
    import pyautogui  # type: ignore
    pyautogui.PAUSE = 0.18
except Exception:
    pyautogui = None


class StopRequested(Exception):
    """Raised when the user presses Stop during the worker loop."""


def enable_dpi_awareness() -> None:
    if os.name != "nt":
        return
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
        except Exception:
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass


enable_dpi_awareness()

DEFAULT_PROFILES = {
    "GPT Image": {"target_x": -1, "target_y": -1, "click_sequence": [], "click_step_wait_sec": 1, "delay_sec": 150, "upload_wait_sec": 20, "file_gap_sec": 8, "countdown_sec": 5, "minimize_on_start": True, "manual_focus": True, "click_before_each_step": True},
    "FLOW": {"target_x": -1, "target_y": -1, "click_sequence": [], "click_step_wait_sec": 1, "delay_sec": 420, "upload_wait_sec": 40, "file_gap_sec": 8, "countdown_sec": 5, "minimize_on_start": True, "manual_focus": True, "click_before_each_step": True},
    "Grok": {"target_x": -1, "target_y": -1, "click_sequence": [], "click_step_wait_sec": 1, "delay_sec": 360, "upload_wait_sec": 35, "file_gap_sec": 8, "countdown_sec": 5, "minimize_on_start": True, "manual_focus": True, "click_before_each_step": True},
    "Custom": {"target_x": -1, "target_y": -1, "click_sequence": [], "click_step_wait_sec": 1, "delay_sec": 180, "upload_wait_sec": 25, "file_gap_sec": 8, "countdown_sec": 5, "minimize_on_start": True, "manual_focus": True, "click_before_each_step": True},
}


def split_paths(value: object) -> list[str]:
    return [p.strip().strip('"') for p in str(value or "").split("|") if p and p.strip()]


def append_unique(target: list[str], paths: Iterable[str]) -> None:
    for path in paths:
        text = str(path or "").strip()
        if text and text not in target:
            target.append(text)


def get_first(row: dict[str, object], *names: str, default: object = "") -> object:
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return default


HEADER_ALIASES = {
    "enable": "enabled", "use": "enabled", "active": "enabled",
    "order": "shot_order", "shotorder": "shot_order", "no": "shot_order", "ลำดับ": "shot_order",
    "shot": "shot_id", "shotid": "shot_id", "id": "shot_id",
    "type": "shot_type", "job_type": "shot_type", "jobtype": "shot_type",
    "category": "category_th", "categoryth": "category_th", "หมวด": "category_th", "หมวดสินค้า": "category_th",
    "scene": "scene_preset", "scene_custom": "scene_custom", "custom_scene": "scene_custom",
    "style": "style_preset",
    "ratio": "aspect_ratio", "aspect": "aspect_ratio", "aspectratio": "aspect_ratio",
    "stillprofile": "profile_still", "profile": "profile_still", "profileimage": "profile_still", "profile_still": "profile_still",
    "motionprofile": "profile_motion", "profilevideo": "profile_motion", "profile_motion": "profile_motion",
    "prompt": "still_prompt", "image_prompt": "still_prompt", "stillprompt": "still_prompt", "first_frame_prompt": "still_prompt",
    "motionprompt": "motion_prompt", "video_prompt": "motion_prompt", "flow_prompt": "motion_prompt",
    "attachment": "attachments", "refs": "attachments", "ref": "attachments", "files": "attachments",
    "productrefs": "product_ref_names", "product_ref": "product_ref_names", "product_ref_name": "product_ref_names",
    "facerefs": "face_ref_names", "face_ref": "face_ref_names", "model_ref": "face_ref_names",
    "outfitrefs": "outfit_ref_names", "outfit_ref": "outfit_ref_names",
    "stylerefs": "style_ref_names", "style_ref": "style_ref_names", "scene_ref": "style_ref_names",
    "delay": "still_delay_sec", "delay_sec": "still_delay_sec", "stilldelay": "still_delay_sec",
    "motiondelay": "motion_delay_sec", "video_delay": "motion_delay_sec",
    "output": "still_output_name", "output_name": "still_output_name", "still_output": "still_output_name",
    "firstframepath": "first_frame_path", "first_frame": "first_frame_path", "first_frame_file": "first_frame_path",
    "motionoutput": "motion_output_name", "video_output": "motion_output_name",
    "status": "still_status", "stillstatus": "still_status", "motionstatus": "motion_status",
    "note": "notes", "remark": "notes", "หมายเหตุ": "notes",
}


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lstrip("\ufeff")
    key = text.lower().replace(" ", "_").replace("-", "_").replace(".", "_")
    compact = key.replace("_", "")
    if key in PROJECT_FIELDNAMES:
        return key
    if compact in HEADER_ALIASES:
        return HEADER_ALIASES[compact]
    if key in HEADER_ALIASES:
        return HEADER_ALIASES[key]
    return key


def is_probable_header(values: Iterable[object]) -> bool:
    headers = [normalize_header(v) for v in values if str(v or "").strip()]
    known = set(headers) & set(PROJECT_FIELDNAMES)
    # Accept either new project format or legacy prompt-sender format.
    return ("shot_id" in known and ("still_prompt" in known or "motion_prompt" in known or "attachments" in known)) or len(known) >= 3


def choose_excel_sheet(workbook):
    for name in ("Project Queue", "Prompt Queue", "Queue", "Sheet1"):
        if name in workbook.sheetnames:
            return workbook[name]
    return workbook.active


def rows_to_items_from_values(rows: list[tuple], default_delay: int, active_profile: str, reset_status: bool) -> list[QueueItem]:
    if not rows:
        return []
    header_idx = None
    for idx, values in enumerate(rows[:30]):
        if is_probable_header(values):
            header_idx = idx
            break
    if header_idx is None:
        # Fall back to first non-empty row to produce a clear import result instead of silently failing.
        for idx, values in enumerate(rows[:30]):
            if any(str(v or "").strip() for v in values):
                header_idx = idx
                break
    if header_idx is None:
        return []
    headers_raw = list(rows[header_idx])
    headers = [normalize_header(value) for value in headers_raw]
    loaded: list[QueueItem] = []
    for values in rows[header_idx + 1:]:
        if not any(str(v or "").strip() for v in values):
            continue
        row = {headers[idx]: value for idx, value in enumerate(values) if idx < len(headers) and headers[idx]}
        item = row_to_item(row, default_delay, active_profile, reset_status=reset_status)
        if item.shot_id or item.still_prompt or item.motion_prompt:
            loaded.append(item)
    return loaded


@dataclass
class QueueItem:
    enabled: str = "yes"
    shot_order: str = ""
    shot_id: str = ""
    shot_type: str = ""
    category_th: str = ""
    scene_preset: str = ""
    scene_custom: str = ""
    style_preset: str = ""
    aspect_ratio: str = "9:16"
    profile_still: str = "GPT Image"
    profile_motion: str = "FLOW"
    still_prompt: str = ""
    motion_prompt: str = ""
    attachments: list[str] = field(default_factory=list)
    product_ref_names: str = ""
    face_ref_names: str = ""
    outfit_ref_names: str = ""
    style_ref_names: str = ""
    still_delay_sec: int = 150
    motion_delay_sec: int = 420
    still_output_name: str = ""
    first_frame_path: str = ""
    motion_output_name: str = ""
    still_status: str = "pending"
    motion_status: str = "pending"
    notes: str = ""

    def active_prompt(self, phase: str) -> str:
        return self.motion_prompt if phase == "motion" else self.still_prompt

    def active_status(self, phase: str) -> str:
        return self.motion_status if phase == "motion" else self.still_status

    def set_active_status(self, phase: str, status: str) -> None:
        if phase == "motion":
            self.motion_status = status
        else:
            self.still_status = status

    def active_delay(self, phase: str) -> int:
        return self.motion_delay_sec if phase == "motion" else self.still_delay_sec

    def active_profile(self, phase: str) -> str:
        return self.profile_motion if phase == "motion" else self.profile_still


def row_to_item(row: dict[str, object], default_delay: int, active_profile: str = "", reset_status: bool = True) -> QueueItem:
    # New project queue format
    if "still_prompt" in row or "motion_prompt" in row or "profile_still" in row or "profile_motion" in row:
        try:
            still_delay = max(1, int(float(str(get_first(row, "still_delay_sec", "delay_sec", default=default_delay)).strip() or default_delay)))
        except ValueError:
            still_delay = default_delay
        try:
            motion_delay = max(1, int(float(str(get_first(row, "motion_delay_sec", "delay_sec", default=420)).strip() or 420)))
        except ValueError:
            motion_delay = 420
        return QueueItem(
            enabled=str(get_first(row, "enabled", default="yes") or "yes").strip() or "yes",
            shot_order=str(get_first(row, "shot_order", default="") or "").strip(),
            shot_id=str(get_first(row, "shot_id", "shot", default="") or "").strip(),
            shot_type=str(get_first(row, "shot_type", "job_type", "type", default="") or "").strip(),
            category_th=str(get_first(row, "category_th", default="") or "").strip(),
            scene_preset=str(get_first(row, "scene_preset", default="") or "").strip(),
            scene_custom=str(get_first(row, "scene_custom", default="") or "").strip(),
            style_preset=str(get_first(row, "style_preset", default="") or "").strip(),
            aspect_ratio=str(get_first(row, "aspect_ratio", default="9:16") or "9:16").strip(),
            profile_still=str(get_first(row, "profile_still", "profile", default=active_profile or "GPT Image") or active_profile or "GPT Image").strip(),
            profile_motion=str(get_first(row, "profile_motion", default="FLOW") or "FLOW").strip(),
            still_prompt=str(get_first(row, "still_prompt", "prompt", default="") or "").strip(),
            motion_prompt=str(get_first(row, "motion_prompt", default="") or "").strip(),
            attachments=split_paths(get_first(row, "attachments", "attachment", default="")),
            product_ref_names=str(get_first(row, "product_ref_names", default="") or "").strip(),
            face_ref_names=str(get_first(row, "face_ref_names", default="") or "").strip(),
            outfit_ref_names=str(get_first(row, "outfit_ref_names", default="") or "").strip(),
            style_ref_names=str(get_first(row, "style_ref_names", default="") or "").strip(),
            still_delay_sec=still_delay,
            motion_delay_sec=motion_delay,
            still_output_name=str(get_first(row, "still_output_name", "output_name", default="") or "").strip(),
            first_frame_path=str(get_first(row, "first_frame_path", default="") or "").strip(),
            motion_output_name=str(get_first(row, "motion_output_name", default="") or "").strip(),
            still_status="pending" if reset_status else (str(get_first(row, "still_status", "status", default="pending") or "pending").strip() or "pending"),
            motion_status="pending" if reset_status else (str(get_first(row, "motion_status", default="pending") or "pending").strip() or "pending"),
            notes=str(get_first(row, "notes", default="") or "").strip(),
        )

    # Legacy queue format
    job_type = str(get_first(row, "job_type", "type", default="first_frame") or "first_frame").strip()
    try:
        delay = max(1, int(float(str(get_first(row, "delay_sec", "delay", default=default_delay)).strip() or default_delay)))
    except ValueError:
        delay = default_delay
    prompt = str(get_first(row, "prompt", default="") or "").strip()
    status = "pending" if reset_status else (str(get_first(row, "status", default="pending") or "pending").strip() or "pending")
    profile = str(get_first(row, "profile", default=active_profile or "GPT Image") or active_profile or "GPT Image").strip()
    is_motion = job_type.lower() in {"motion", "video", "flow"}
    return QueueItem(
        shot_id=str(get_first(row, "shot_id", "shot", default="") or "").strip(),
        shot_type=job_type,
        profile_still=profile if not is_motion else "GPT Image",
        profile_motion=profile if is_motion else "FLOW",
        still_prompt="" if is_motion else prompt,
        motion_prompt=prompt if is_motion else "",
        attachments=split_paths(get_first(row, "attachments", "attachment", default="")),
        still_delay_sec=delay if not is_motion else default_delay,
        motion_delay_sec=delay if is_motion else 420,
        still_output_name=str(get_first(row, "output_name", default="") or "").strip() if not is_motion else "",
        motion_output_name=str(get_first(row, "output_name", default="") or "").strip() if is_motion else "",
        first_frame_path="" if not is_motion else "|".join(split_paths(get_first(row, "attachments", "attachment", default=""))),
        still_status=status if not is_motion else "pending",
        motion_status=status if is_motion else "pending",
    )


def item_to_row(item: QueueItem) -> dict[str, object]:
    return {
        "enabled": item.enabled,
        "shot_order": item.shot_order,
        "shot_id": item.shot_id,
        "shot_type": item.shot_type,
        "category_th": item.category_th,
        "scene_preset": item.scene_preset,
        "scene_custom": item.scene_custom,
        "style_preset": item.style_preset,
        "aspect_ratio": item.aspect_ratio,
        "profile_still": item.profile_still,
        "profile_motion": item.profile_motion,
        "still_prompt": item.still_prompt,
        "motion_prompt": item.motion_prompt,
        "attachments": "|".join(item.attachments),
        "product_ref_names": item.product_ref_names,
        "face_ref_names": item.face_ref_names,
        "outfit_ref_names": item.outfit_ref_names,
        "style_ref_names": item.style_ref_names,
        "still_delay_sec": item.still_delay_sec,
        "motion_delay_sec": item.motion_delay_sec,
        "still_output_name": item.still_output_name,
        "first_frame_path": item.first_frame_path,
        "motion_output_name": item.motion_output_name,
        "still_status": item.still_status,
        "motion_status": item.motion_status,
        "notes": item.notes,
    }


def load_items_from_file(path: str, default_delay: int, active_profile: str = "", reset_status: bool = True) -> list[QueueItem]:
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        raise RuntimeError("ไฟล์ .xls รุ่นเก่ายังไม่รองรับค่ะ กรุณา Save As เป็น .xlsx หรือ CSV UTF-8 ก่อน import")
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("Excel import ต้องใช้ openpyxl ค่ะ ให้รัน install.bat หรือสั่ง: py -m pip install openpyxl") from exc
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = choose_excel_sheet(workbook)
            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()
        except Exception as exc:
            raise RuntimeError(f"เปิดไฟล์ Excel ไม่สำเร็จ: {exc}") from exc
        loaded = rows_to_items_from_values(rows, default_delay, active_profile, reset_status)
        if not loaded:
            raise RuntimeError("อ่าน Excel แล้วไม่พบแถวงาน ตรวจว่ามี header เช่น shot_id, still_prompt, motion_prompt หรือใช้ template project_queue_template.xlsx")
        return loaded
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return []
        normalized = [normalize_header(h) for h in reader.fieldnames]
        loaded: list[QueueItem] = []
        for raw in reader:
            row = {normalized[idx]: raw.get(reader.fieldnames[idx], "") for idx in range(len(reader.fieldnames))}
            item = row_to_item(row, default_delay, active_profile, reset_status=reset_status)
            if item.shot_id or item.still_prompt or item.motion_prompt:
                loaded.append(item)
        return loaded


def save_items_to_file(path: str, items: list[QueueItem]) -> None:
    suffix = Path(path).suffix.lower()
    rows = [item_to_row(item) for item in items]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import Workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("Excel export needs openpyxl. Run install.bat first.") from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Project Queue"
        sheet.append(PROJECT_FIELDNAMES)
        for row in rows:
            sheet.append([row.get(name, "") for name in PROJECT_FIELDNAMES])
        for column in ("A", "B", "C", "D", "K", "L", "M", "N", "U", "V", "W", "X", "Y"):
            sheet.column_dimensions[column].width = 18
        sheet.column_dimensions["L"].width = 82
        sheet.column_dimensions["M"].width = 82
        workbook.save(path)
        workbook.close()
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=PROJECT_FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def retry(label: str, func, attempts: int = 3, delay: float = 0.35):
    last: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            return func()
        except Exception as exc:
            last = exc
            if attempt == attempts:
                break
            time.sleep(delay * attempt)
    raise RuntimeError(f"{label} failed: {last}") from last


def set_text_clipboard_windows(text: str) -> None:
    if os.name != "nt":
        raise RuntimeError("Text clipboard native mode is Windows-only.")
    import ctypes
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    kernel32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
    kernel32.GlobalAlloc.restype = ctypes.c_void_p
    kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
    kernel32.GlobalLock.restype = ctypes.c_void_p
    kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]
    kernel32.GlobalUnlock.restype = ctypes.c_bool
    kernel32.GlobalFree.argtypes = [ctypes.c_void_p]
    kernel32.GlobalFree.restype = ctypes.c_void_p
    user32.OpenClipboard.argtypes = [ctypes.c_void_p]
    user32.OpenClipboard.restype = ctypes.c_bool
    user32.EmptyClipboard.argtypes = []
    user32.EmptyClipboard.restype = ctypes.c_bool
    user32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
    user32.SetClipboardData.restype = ctypes.c_void_p
    user32.CloseClipboard.argtypes = []
    user32.CloseClipboard.restype = ctypes.c_bool
    payload = (text + "\0").encode("utf-16le")
    handle = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(payload))
    if not handle:
        raise OSError("GlobalAlloc failed for text")
    locked = kernel32.GlobalLock(handle)
    if not locked:
        kernel32.GlobalFree(handle)
        raise OSError("GlobalLock failed for text")
    try:
        ctypes.memmove(locked, payload, len(payload))
    finally:
        kernel32.GlobalUnlock(handle)
    if not user32.OpenClipboard(None):
        kernel32.GlobalFree(handle)
        raise OSError("OpenClipboard failed for text")
    try:
        user32.EmptyClipboard()
        if not user32.SetClipboardData(CF_UNICODETEXT, handle):
            kernel32.GlobalFree(handle)
            raise OSError("SetClipboardData failed for text")
        handle = None
    finally:
        user32.CloseClipboard()


def set_file_clipboard_windows(paths: Iterable[str]) -> None:
    if os.name != "nt":
        raise RuntimeError("File clipboard mode is Windows-only.")
    import ctypes
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    kernel32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
    kernel32.GlobalAlloc.restype = ctypes.c_void_p
    kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
    kernel32.GlobalLock.restype = ctypes.c_void_p
    kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]
    kernel32.GlobalUnlock.restype = ctypes.c_bool
    kernel32.GlobalFree.argtypes = [ctypes.c_void_p]
    kernel32.GlobalFree.restype = ctypes.c_void_p
    user32.OpenClipboard.argtypes = [ctypes.c_void_p]
    user32.OpenClipboard.restype = ctypes.c_bool
    user32.EmptyClipboard.argtypes = []
    user32.EmptyClipboard.restype = ctypes.c_bool
    user32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
    user32.SetClipboardData.restype = ctypes.c_void_p
    user32.CloseClipboard.argtypes = []
    user32.CloseClipboard.restype = ctypes.c_bool
    absolute = [str(Path(p).resolve()) for p in paths if p]
    missing = [p for p in absolute if not Path(p).exists()]
    if missing:
        raise FileNotFoundError("Missing attachment: " + missing[0])
    if not absolute:
        return
    encoded = ("\0".join(absolute) + "\0\0").encode("utf-16le")
    payload = struct.pack("IiiII", 20, 0, 0, 0, 1) + encoded
    handle = kernel32.GlobalAlloc(GMEM_MOVEABLE, len(payload))
    if not handle:
        raise OSError("GlobalAlloc failed for files")
    locked = kernel32.GlobalLock(handle)
    if not locked:
        kernel32.GlobalFree(handle)
        raise OSError("GlobalLock failed for files")
    try:
        ctypes.memmove(locked, payload, len(payload))
    finally:
        kernel32.GlobalUnlock(handle)
    if not user32.OpenClipboard(None):
        kernel32.GlobalFree(handle)
        raise OSError("OpenClipboard failed for files")
    try:
        user32.EmptyClipboard()
        if not user32.SetClipboardData(CF_HDROP, handle):
            kernel32.GlobalFree(handle)
            raise OSError("SetClipboardData failed for files")
        handle = None
    finally:
        user32.CloseClipboard()


def send_ctrl_v() -> None:
    if pyautogui is None:
        raise RuntimeError("pyautogui is not installed.")
    if os.name != "nt":
        pyautogui.hotkey("ctrl", "v")
        return
    import ctypes
    user32 = ctypes.windll.user32
    keybd_event = user32.keybd_event
    keybd_event.argtypes = [ctypes.c_ubyte, ctypes.c_ubyte, ctypes.c_uint, ctypes.c_void_p]
    keybd_event.restype = None
    vk_control = 0x11
    vk_v = 0x56
    key_up = 0x0002
    keybd_event(vk_control, 0, 0, None)
    time.sleep(0.04)
    keybd_event(vk_v, 0, 0, None)
    time.sleep(0.04)
    keybd_event(vk_v, 0, key_up, None)
    time.sleep(0.04)
    keybd_event(vk_control, 0, key_up, None)


def send_enter() -> None:
    if pyautogui is None:
        raise RuntimeError("pyautogui is not installed.")
    if os.name != "nt":
        pyautogui.press("enter")
        return
    import ctypes
    user32 = ctypes.windll.user32
    keybd_event = user32.keybd_event
    keybd_event.argtypes = [ctypes.c_ubyte, ctypes.c_ubyte, ctypes.c_uint, ctypes.c_void_p]
    keybd_event.restype = None
    vk_return = 0x0D
    key_up = 0x0002
    keybd_event(vk_return, 0, 0, None)
    time.sleep(0.05)
    keybd_event(vk_return, 0, key_up, None)


def native_click(x: int, y: int) -> None:
    if pyautogui is None:
        raise RuntimeError("pyautogui is not installed.")
    if os.name != "nt":
        pyautogui.click(x, y)
        return
    import ctypes
    user32 = ctypes.windll.user32
    user32.SetCursorPos(int(x), int(y))
    time.sleep(0.18)
    user32.mouse_event(0x0002, 0, 0, 0, None)
    time.sleep(0.05)
    user32.mouse_event(0x0004, 0, 0, 0, None)


class PromptSenderApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1060x760")
        self.minsize(920, 640)
        self.configure(bg="#0b1017")
        self.items: list[QueueItem] = []
        self.queue_path: str | None = None
        self.profiles = self.load_profiles()
        self.ui_queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.stop_event = threading.Event()
        self.worker: threading.Thread | None = None
        self.busy = False
        self.global_refs: dict[str, list[str]] = {"product": [], "face": [], "outfit": [], "style": []}

        self.profile_var = tk.StringVar(value=self.load_last_profile())
        self.phase_var = tk.StringVar(value="still")
        self.target_x_var = tk.IntVar(value=-1)
        self.target_y_var = tk.IntVar(value=-1)
        self.click_steps: list[dict[str, int]] = []
        self.click_step_wait_var = tk.IntVar(value=1)
        self.delay_var = tk.IntVar(value=150)
        self.upload_var = tk.IntVar(value=20)
        self.file_gap_var = tk.IntVar(value=8)
        self.countdown_var = tk.IntVar(value=5)
        self.minimize_var = tk.BooleanVar(value=True)
        self.manual_focus_var = tk.BooleanVar(value=True)
        self.click_each_var = tk.BooleanVar(value=True)
        self.skip_sent_var = tk.BooleanVar(value=True)
        self.autosave_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Ready. Import Project Queue, choose Still/Motion phase, then Start.")
        self.attachment_count_var = tk.StringVar(value="Refs ready: 0")
        self.global_ref_list: tk.Listbox | None = None
        self.click_steps_list: tk.Listbox | None = None

        self.configure_style()
        self.build_ui()
        self.apply_profile_to_ui(self.profile_var.get())
        self.after(200, self.drain_ui_queue)

    def configure_style(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", background="#0b1017", foreground="#d7e3f2", fieldbackground="#111827")
        style.configure("Panel.TFrame", background="#0b1017")
        style.configure("Card.TFrame", background="#111827")
        style.configure("Title.TLabel", background="#0b1017", foreground="#f8fafc", font=("Segoe UI", 16, "bold"))
        style.configure("Subtle.TLabel", background="#0b1017", foreground="#91a4bd", font=("Segoe UI", 9))
        style.configure("CardLabel.TLabel", background="#111827", foreground="#93c5fd", font=("Segoe UI", 8, "bold"))
        style.configure("TButton", background="#1f2937", foreground="#e5edf8", borderwidth=0, padding=(10, 7))
        style.configure("Accent.TButton", background="#2dd4bf", foreground="#04111a", font=("Segoe UI", 9, "bold"), padding=(10, 9))
        style.configure("Danger.TButton", background="#7f1d1d", foreground="#fee2e2", padding=(10, 9))
        style.configure("Treeview", background="#0f1724", foreground="#e5edf8", fieldbackground="#0f1724", rowheight=28)
        style.configure("Treeview.Heading", background="#162033", foreground="#93c5fd", font=("Segoe UI", 8, "bold"))
        style.map("Treeview", background=[("selected", "#0f766e")], foreground=[("selected", "#ffffff")])
        style.configure("TCheckbutton", background="#111827", foreground="#cbd5e1")

    def build_ui(self) -> None:
        root = ttk.Frame(self, style="Panel.TFrame", padding=12)
        root.pack(fill="both", expand=True)
        header = ttk.Frame(root, style="Panel.TFrame")
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text="WePlus Prompt Sender", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header, text="Project Queue: one row contains still + motion prompts. Last click step is the paste target.", style="Subtle.TLabel").pack(anchor="w")

        top = ttk.Frame(root, style="Panel.TFrame")
        top.pack(fill="x", pady=(0, 10))
        top.columnconfigure(0, weight=1)
        top.columnconfigure(1, weight=1)

        profile_card = ttk.Frame(top, style="Card.TFrame", padding=10)
        profile_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ttk.Label(profile_card, text="PROFILE / CLICK BEHAVIOR", style="CardLabel.TLabel").grid(row=0, column=0, columnspan=4, sticky="w")
        self.profile_combo = ttk.Combobox(profile_card, textvariable=self.profile_var, values=sorted(self.profiles.keys()), state="readonly")
        self.profile_combo.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 6))
        self.profile_combo.bind("<<ComboboxSelected>>", lambda _e: self.apply_profile_to_ui(self.profile_var.get()))
        ttk.Button(profile_card, text="Save Profile", command=self.save_current_profile).grid(row=1, column=2, sticky="ew", padx=(6, 0), pady=(8, 6))
        ttk.Button(profile_card, text="New Profile", command=self.new_profile).grid(row=1, column=3, sticky="ew", padx=(6, 0), pady=(8, 6))
        ttk.Label(profile_card, text="Target X", style="CardLabel.TLabel").grid(row=2, column=0, sticky="w")
        ttk.Entry(profile_card, textvariable=self.target_x_var, width=8).grid(row=2, column=1, sticky="ew")
        ttk.Label(profile_card, text="Target Y", style="CardLabel.TLabel").grid(row=2, column=2, sticky="w", padx=(10, 0))
        ttk.Entry(profile_card, textvariable=self.target_y_var, width=8).grid(row=2, column=3, sticky="ew")
        ttk.Button(profile_card, text="Set Final Point", command=self.set_point).grid(row=3, column=0, columnspan=2, sticky="ew", pady=(8, 0))
        ttk.Button(profile_card, text="Test Final Click", command=self.test_click).grid(row=3, column=2, columnspan=2, sticky="ew", padx=(6, 0), pady=(8, 0))
        ttk.Label(profile_card, text="Multi-click sequence: จุดสุดท้ายคือจุดวางภาพ+prompt", style="CardLabel.TLabel").grid(row=4, column=0, columnspan=4, sticky="w", pady=(10, 0))
        self.click_steps_list = tk.Listbox(profile_card, height=3, bg="#0b1220", fg="#e5edf8", relief="flat")
        self.click_steps_list.grid(row=5, column=0, columnspan=4, sticky="ew", pady=(4, 4))
        ttk.Button(profile_card, text="Add Click Step", command=self.add_click_step).grid(row=6, column=0, sticky="ew", pady=(4, 0))
        ttk.Button(profile_card, text="Remove Step", command=self.remove_click_step).grid(row=6, column=1, sticky="ew", padx=(6, 0), pady=(4, 0))
        ttk.Button(profile_card, text="Clear Steps", command=self.clear_click_steps).grid(row=6, column=2, sticky="ew", padx=(6, 0), pady=(4, 0))
        ttk.Button(profile_card, text="Test Sequence", command=self.test_click_sequence).grid(row=6, column=3, sticky="ew", padx=(6, 0), pady=(4, 0))
        ttk.Label(profile_card, text="Step wait", style="CardLabel.TLabel").grid(row=7, column=0, sticky="w", pady=(6, 0))
        ttk.Spinbox(profile_card, from_=0, to=30, textvariable=self.click_step_wait_var, width=8).grid(row=7, column=1, sticky="ew", pady=(6, 0))
        for c in range(4):
            profile_card.columnconfigure(c, weight=1)

        timing_card = ttk.Frame(top, style="Card.TFrame", padding=10)
        timing_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        ttk.Label(timing_card, text="PHASE / TIMING", style="CardLabel.TLabel").grid(row=0, column=0, columnspan=4, sticky="w")
        ttk.Label(timing_card, text="Phase", style="CardLabel.TLabel").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.phase_combo = ttk.Combobox(timing_card, textvariable=self.phase_var, values=["still", "motion"], state="readonly", width=10)
        self.phase_combo.grid(row=1, column=1, sticky="ew", padx=(5, 0), pady=(8, 0))
        self.phase_combo.bind("<<ComboboxSelected>>", lambda _e: self.refresh_table())
        self.spin(timing_card, 1, 2, "Delay", self.delay_var, 1, 1800)
        self.spin(timing_card, 2, 0, "Upload", self.upload_var, 1, 300)
        self.spin(timing_card, 2, 2, "File gap", self.file_gap_var, 1, 120)
        self.spin(timing_card, 3, 0, "Countdown", self.countdown_var, 1, 60)
        ttk.Button(timing_card, text="Test Paste Prompt", command=self.test_paste_prompt).grid(row=3, column=2, sticky="ew", padx=(6, 0), pady=(8, 0))
        ttk.Button(timing_card, text="STOP", style="Danger.TButton", command=self.stop).grid(row=3, column=3, sticky="ew", padx=(6, 0), pady=(8, 0))
        ttk.Checkbutton(timing_card, text="Minimize on start/countdown", variable=self.minimize_var).grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Checkbutton(timing_card, text="Manual focus countdown", variable=self.manual_focus_var).grid(row=4, column=2, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Checkbutton(timing_card, text="Click before each paste", variable=self.click_each_var).grid(row=5, column=0, columnspan=2, sticky="w")
        ttk.Checkbutton(timing_card, text="Skip sent rows in active phase", variable=self.skip_sent_var).grid(row=5, column=2, columnspan=2, sticky="w")
        ttk.Checkbutton(timing_card, text="Autosave status to *.autosave.csv", variable=self.autosave_var).grid(row=6, column=0, columnspan=4, sticky="w")
        for c in range(4):
            timing_card.columnconfigure(c, weight=1)

        refs_card = ttk.Frame(root, style="Card.TFrame", padding=10)
        refs_card.pack(fill="x", pady=(0, 10))
        ttk.Label(refs_card, text="GLOBAL REFS ใช้ซ้ำทั้งโปรเจกต์", style="CardLabel.TLabel").grid(row=0, column=0, columnspan=5, sticky="w")
        ttk.Button(refs_card, text="Add Product", command=lambda: self.add_global_refs("product")).grid(row=1, column=0, sticky="ew", padx=(0, 5), pady=(8, 0))
        ttk.Button(refs_card, text="Add Face", command=lambda: self.add_global_refs("face")).grid(row=1, column=1, sticky="ew", padx=(0, 5), pady=(8, 0))
        ttk.Button(refs_card, text="Add Outfit", command=lambda: self.add_global_refs("outfit")).grid(row=1, column=2, sticky="ew", padx=(0, 5), pady=(8, 0))
        ttk.Button(refs_card, text="Add Style", command=lambda: self.add_global_refs("style")).grid(row=1, column=3, sticky="ew", padx=(0, 5), pady=(8, 0))
        ttk.Button(refs_card, text="Clear Refs", command=self.clear_global_refs).grid(row=1, column=4, sticky="ew", pady=(8, 0))
        self.global_ref_list = tk.Listbox(refs_card, height=3, bg="#0b1220", fg="#e5edf8", relief="flat")
        self.global_ref_list.grid(row=2, column=0, columnspan=5, sticky="ew", pady=(8, 0))
        ttk.Label(refs_card, textvariable=self.attachment_count_var, style="CardLabel.TLabel").grid(row=3, column=0, columnspan=5, sticky="w", pady=(6, 0))
        for c in range(5):
            refs_card.columnconfigure(c, weight=1)

        action = ttk.Frame(root, style="Panel.TFrame")
        action.pack(fill="x", pady=(0, 6))
        ttk.Button(action, text="Replace Project CSV", command=self.import_queue).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Append CSV", command=self.append_queue).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Export Queue", command=self.export_queue).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Add", command=self.add_item).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Edit", command=self.edit_item).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Duplicate", command=self.duplicate_item).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Remove", command=self.remove_item).pack(side="left", padx=(0, 6))
        ttk.Button(action, text="Start", style="Accent.TButton", command=self.start).pack(side="right", padx=(6, 0))
        ttk.Button(action, text="Stop", style="Danger.TButton", command=self.stop).pack(side="right")

        action2 = ttk.Frame(root, style="Panel.TFrame")
        action2.pack(fill="x", pady=(0, 10))
        ttk.Button(action2, text="Send Selected", command=self.send_selected).pack(side="left", padx=(0, 6))
        ttk.Button(action2, text="Reset Selected", command=self.reset_selected_pending).pack(side="left", padx=(0, 6))
        ttk.Button(action2, text="Reset All", command=self.reset_all_pending).pack(side="left", padx=(0, 6))
        ttk.Button(action2, text="Delete Sent", command=self.delete_sent).pack(side="left", padx=(0, 6))
        ttk.Button(action2, text="Clear Queue", command=self.clear_queue).pack(side="left", padx=(0, 6))

        table_card = ttk.Frame(root, style="Card.TFrame", padding=10)
        table_card.pack(fill="both", expand=True, pady=(0, 8))
        columns = ("enabled", "shot", "type", "still", "motion", "first_frame", "refs", "sec", "output")
        self.tree = ttk.Treeview(table_card, columns=columns, show="headings", selectmode="extended")
        headings = {"enabled":"On", "shot":"Shot", "type":"Type", "still":"Still", "motion":"Motion", "first_frame":"First Frame Path", "refs":"Refs", "sec":"Sec", "output":"Output"}
        widths = {"enabled":45, "shot":110, "type":150, "still":80, "motion":80, "first_frame":210, "refs":55, "sec":45, "output":140}
        for c in columns:
            self.tree.heading(c, text=headings[c])
            self.tree.column(c, width=widths[c], anchor="center" if c in {"enabled", "still", "motion", "refs", "sec"} else "w")
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", lambda _e: self.refresh_attachment_count())
        scroll = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
        scroll.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scroll.set)
        ttk.Label(root, textvariable=self.status_var, style="Subtle.TLabel", wraplength=1020).pack(anchor="w", fill="x")

    def spin(self, parent, row, col, label, var, min_v, max_v) -> None:
        ttk.Label(parent, text=label, style="CardLabel.TLabel").grid(row=row, column=col, sticky="w", pady=(8, 0))
        ttk.Spinbox(parent, from_=min_v, to=max_v, textvariable=var, width=8).grid(row=row, column=col+1, sticky="ew", padx=(5, 0), pady=(8, 0))

    def load_profiles(self) -> dict:
        if PROFILE_PATH.exists():
            try:
                data = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
                profiles = dict(DEFAULT_PROFILES)
                profiles.update(data.get("profiles", {}))
                return profiles
            except Exception:
                pass
        return dict(DEFAULT_PROFILES)

    def load_last_profile(self) -> str:
        if PROFILE_PATH.exists():
            try:
                data = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
                last = data.get("last_profile")
                if last in self.profiles:
                    return last
            except Exception:
                pass
        return "GPT Image"

    def persist_profiles(self) -> None:
        PROFILE_PATH.write_text(json.dumps({"last_profile": self.profile_var.get(), "profiles": self.profiles}, ensure_ascii=False, indent=2), encoding="utf-8")

    def current_profile_data(self) -> dict:
        return {"target_x": int(self.target_x_var.get()), "target_y": int(self.target_y_var.get()), "click_sequence": list(self.click_steps), "click_step_wait_sec": int(self.click_step_wait_var.get()), "delay_sec": int(self.delay_var.get()), "upload_wait_sec": int(self.upload_var.get()), "file_gap_sec": int(self.file_gap_var.get()), "countdown_sec": int(self.countdown_var.get()), "minimize_on_start": bool(self.minimize_var.get()), "manual_focus": bool(self.manual_focus_var.get()), "click_before_each_step": bool(self.click_each_var.get())}

    def apply_profile_to_ui(self, name: str) -> None:
        data = self.profiles.get(name) or DEFAULT_PROFILES["Custom"]
        self.target_x_var.set(int(data.get("target_x", -1)))
        self.target_y_var.set(int(data.get("target_y", -1)))
        self.click_steps = []
        for step in data.get("click_sequence") or []:
            try:
                self.click_steps.append({"x": int(step.get("x", -1)), "y": int(step.get("y", -1)), "wait": int(step.get("wait", data.get("click_step_wait_sec", 1)))})
            except Exception:
                pass
        self.click_step_wait_var.set(int(data.get("click_step_wait_sec", 1)))
        self.refresh_click_steps_list()
        self.delay_var.set(int(data.get("delay_sec", 150)))
        self.upload_var.set(int(data.get("upload_wait_sec", 20)))
        self.file_gap_var.set(int(data.get("file_gap_sec", 8)))
        self.countdown_var.set(int(data.get("countdown_sec", 5)))
        self.minimize_var.set(bool(data.get("minimize_on_start", True)))
        self.manual_focus_var.set(bool(data.get("manual_focus", True)))
        self.click_each_var.set(bool(data.get("click_before_each_step", True)))
        self.persist_profiles()
        self.status_var.set(f"Loaded profile: {name}")

    def save_current_profile(self) -> None:
        name = self.profile_var.get().strip() or "Custom"
        self.profiles[name] = self.current_profile_data()
        self.profile_combo.configure(values=sorted(self.profiles.keys()))
        self.persist_profiles()
        self.status_var.set(f"Saved profile: {name}")

    def new_profile(self) -> None:
        name = simpledialog.askstring(APP_TITLE, "New profile name:")
        if not name:
            return
        self.profiles[name] = self.current_profile_data()
        self.profile_var.set(name)
        self.profile_combo.configure(values=sorted(self.profiles.keys()))
        self.persist_profiles()

    def set_point(self) -> None:
        try:
            x, y = self.capture_mouse_point("final paste point")
        except Exception as exc:
            self.deiconify(); messagebox.showerror(APP_TITLE, str(exc)); return
        self.target_x_var.set(x); self.target_y_var.set(y)
        if self.click_steps:
            self.click_steps[-1] = {"x": x, "y": y, "wait": max(0, int(self.click_step_wait_var.get()))}
            self.refresh_click_steps_list()
        self.save_current_profile()
        self.log(f"Final paste point set: {x}, {y}")

    def capture_mouse_point(self, label: str) -> tuple[int, int]:
        if pyautogui is None:
            raise RuntimeError("pyautogui is not installed. Run install.bat first.")
        if self.minimize_var.get(): self.iconify()
        for remaining in range(max(1, int(self.countdown_var.get())), 0, -1):
            self.status_var.set(f"Move mouse to {label}. Capturing in {remaining}s...")
            self.update_idletasks(); time.sleep(1)
        pos = pyautogui.position()
        self.deiconify()
        return int(pos.x), int(pos.y)

    def click_target(self) -> bool:
        x, y = int(self.target_x_var.get()), int(self.target_y_var.get())
        if x < 0 or y < 0: return False
        native_click(x, y); return True

    def test_click(self) -> None:
        try: self.log(f"Test click result: {self.click_target()}")
        except Exception as exc: self.log(f"Test click failed: {exc}")

    def add_click_step(self) -> None:
        try:
            x, y = self.capture_mouse_point("next click step")
        except Exception as exc:
            self.deiconify(); messagebox.showerror(APP_TITLE, str(exc)); return
        self.click_steps.append({"x": x, "y": y, "wait": max(0, int(self.click_step_wait_var.get()))})
        self.target_x_var.set(x); self.target_y_var.set(y)
        self.refresh_click_steps_list(); self.save_current_profile()

    def remove_click_step(self) -> None:
        if not self.click_steps: return
        sel = self.click_steps_list.curselection() if self.click_steps_list else ()
        idx = int(sel[0]) if sel else len(self.click_steps) - 1
        if 0 <= idx < len(self.click_steps):
            self.click_steps.pop(idx)
            if self.click_steps:
                self.target_x_var.set(int(self.click_steps[-1].get("x", -1)))
                self.target_y_var.set(int(self.click_steps[-1].get("y", -1)))
            self.refresh_click_steps_list(); self.save_current_profile()

    def clear_click_steps(self) -> None:
        self.click_steps = []
        self.refresh_click_steps_list(); self.save_current_profile()

    def refresh_click_steps_list(self) -> None:
        if not self.click_steps_list: return
        self.click_steps_list.delete(0, "end")
        for idx, step in enumerate(self.click_steps, start=1):
            role = "PASTE HERE" if idx == len(self.click_steps) else "click"
            self.click_steps_list.insert("end", f"{idx}. {role}: x={step.get('x')} y={step.get('y')} wait={step.get('wait', 1)}s")

    def run_click_sequence(self, label: str = "") -> bool:
        if self.click_steps:
            for idx, step in enumerate(self.click_steps, start=1):
                if self.stop_event.is_set():
                    raise StopRequested(STOP_MESSAGE)
                x, y = int(step.get("x", -1)), int(step.get("y", -1))
                if x >= 0 and y >= 0:
                    native_click(x, y)
                wait = max(0, int(step.get("wait", self.click_step_wait_var.get())))
                role = "paste target" if idx == len(self.click_steps) else "step"
                self.write_log(f"{label}: click sequence {idx}/{len(self.click_steps)} ({role}) at {x},{y}; wait={wait}s")
                if wait: time.sleep(wait)
            return True
        return self.click_target()

    def test_click_sequence(self) -> None:
        try: self.log(f"Test sequence result: {self.run_click_sequence('TEST sequence')}. Paste target is last step.")
        except Exception as exc: self.log(f"Test sequence failed: {exc}")

    def focus_countdown(self, label: str) -> None:
        if self.manual_focus_var.get():
            if self.minimize_var.get(): self.iconify()
            for remaining in range(max(1, int(self.countdown_var.get())), 0, -1):
                if self.stop_event.is_set():
                    raise StopRequested(STOP_MESSAGE)
                self.ui_queue.put(("status", f"Click composer for {label}: {remaining}s")); time.sleep(1)

    def test_paste_prompt(self) -> None:
        try:
            self.focus_countdown("TEST")
            if self.click_each_var.get(): self.run_click_sequence("TEST paste")
            set_text_clipboard_windows("TEST PROMPT from WePlus Prompt Sender")
            time.sleep(0.2); send_ctrl_v(); self.deiconify(); self.log("Test prompt pasted.")
        except Exception as exc:
            self.deiconify(); self.log(f"Test prompt failed: {exc}")

    def global_paths(self) -> list[str]:
        paths: list[str] = []
        for kind in ("product", "face", "outfit", "style"):
            append_unique(paths, self.global_refs.get(kind, []))
        return paths

    def final_attachments(self, item: QueueItem, phase: str | None = None) -> list[str]:
        """Return the exact attachment list for one send.

        Still phase intentionally uses either session global refs OR row refs,
        never both. This keeps Add Product/Face/Outfit/Style predictable and
        prevents duplicated pasted images.
        """
        phase = phase or self.phase_var.get()
        paths: list[str] = []
        if phase == "motion":
            append_unique(paths, split_paths(item.first_frame_path))
            # allow fallback for old motion rows if first_frame_path is empty
            if not paths:
                append_unique(paths, item.attachments)
            return paths

        global_paths = self.global_paths()
        if global_paths:
            append_unique(paths, global_paths)
            return paths

        append_unique(paths, item.attachments)
        return paths

    def refresh_global_refs(self) -> None:
        if not self.global_ref_list: return
        self.global_ref_list.delete(0, "end")
        for kind in ("product", "face", "outfit", "style"):
            for path in self.global_refs.get(kind, []):
                self.global_ref_list.insert("end", f"{kind}: {path}")
        self.refresh_table()

    def add_global_refs(self, kind: str) -> None:
        files = filedialog.askopenfilenames(title=f"Select {kind} reference files")
        append_unique(self.global_refs.setdefault(kind, []), files)
        self.refresh_global_refs()
        self.status_var.set(f"Added {len(files)} {kind} global refs.")

    def clear_global_refs(self) -> None:
        self.global_refs = {"product": [], "face": [], "outfit": [], "style": []}
        self.refresh_global_refs(); self.status_var.set("Cleared global refs.")

    def refresh_attachment_count(self) -> None:
        phase = self.phase_var.get()
        selected = self.selected_indices()
        if selected:
            counts = [len(self.final_attachments(self.items[idx], phase)) for idx in selected]
            if len(counts) == 1:
                self.attachment_count_var.set(f"Refs ready for selected: {counts[0]}")
            else:
                self.attachment_count_var.set(f"Refs ready for {len(counts)} selected: {min(counts)}-{max(counts)}")
            return
        if phase == "still":
            self.attachment_count_var.set(f"Refs ready: {len(self.global_paths())} global still refs")
        else:
            motion_ready = sum(1 for item in self.items if self.final_attachments(item, "motion"))
            self.attachment_count_var.set(f"Refs ready: {motion_ready}/{len(self.items)} motion rows have first frame")

    def refresh_table(self) -> None:
        self.tree.delete(*self.tree.get_children())
        phase = self.phase_var.get()
        for idx, item in enumerate(self.items):
            first_path = item.first_frame_path
            if len(first_path) > 34: first_path = "..." + first_path[-31:]
            self.tree.insert("", "end", iid=str(idx), values=(
                item.enabled, item.shot_id, item.shot_type,
                item.still_status, item.motion_status, first_path or "-",
                len(self.final_attachments(item, phase)) or "-", item.active_delay(phase),
                item.motion_output_name if phase == "motion" else item.still_output_name,
            ))
        self.refresh_attachment_count()

    def import_queue(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("CSV / Excel", "*.csv *.xlsx *.xlsm"), ("All files", "*.*")])
        if not path: return
        try:
            self.items = load_items_from_file(path, self.delay_var.get(), self.profile_var.get(), reset_status=True)
            self.queue_path = path
            # Replace means a clean queue. Clear old global refs so old product/face
            # images are not accidentally reused with a new project.
            self.global_refs = {"product": [], "face": [], "outfit": [], "style": []}
            self.refresh_global_refs()
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc)); return
        self.refresh_table(); self.status_var.set(f"Replaced queue with {len(self.items)} fresh pending rows. Global refs cleared: {path}")

    def append_queue(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("CSV / Excel", "*.csv *.xlsx *.xlsm"), ("All files", "*.*")])
        if not path: return
        try:
            incoming = load_items_from_file(path, self.delay_var.get(), self.profile_var.get(), reset_status=True)
            self.items.extend(incoming)
            if not self.queue_path: self.queue_path = path
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc)); return
        self.refresh_table(); self.status_var.set(f"Appended {len(incoming)} fresh pending rows from: {path}")

    def export_queue(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx"), ("All files", "*.*")])
        if not path: return
        try: save_items_to_file(path, self.items)
        except Exception as exc: messagebox.showerror(APP_TITLE, str(exc)); return
        self.queue_path = path; self.status_var.set(f"Saved queue: {path}")

    def autosave_queue(self) -> None:
        if self.autosave_var.get() and self.queue_path:
            try:
                src = Path(self.queue_path)
                save_items_to_file(str(src.with_name(src.stem + ".autosave.csv")), self.items)
            except Exception as exc:
                self.ui_queue.put(("status", f"Autosave failed: {exc}"))

    def selected_index(self) -> int | None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo(APP_TITLE, "Select a row first."); return None
        return int(selected[0])

    def selected_indices(self) -> list[int]:
        return [int(i) for i in self.tree.selection()]

    def duplicate_item(self) -> None:
        idx = self.selected_index()
        if idx is None: return
        src = self.items[idx]
        copy = QueueItem(**item_to_row(src))
        copy.attachments = list(src.attachments)
        copy.shot_id = f"{src.shot_id}_COPY"
        copy.still_status = "pending"; copy.motion_status = "pending"
        self.items.insert(idx + 1, copy)
        self.refresh_table(); self.status_var.set(f"Duplicated {src.shot_id} as pending.")

    def reset_selected_pending(self) -> None:
        indices = self.selected_indices()
        if not indices:
            idx = self.selected_index()
            if idx is None: return
            indices = [idx]
        phase = self.phase_var.get()
        for idx in indices: self.items[idx].set_active_status(phase, "pending")
        self.refresh_table(); self.status_var.set(f"Reset {len(indices)} selected rows to pending in {phase} phase.")

    def reset_all_pending(self) -> None:
        phase = self.phase_var.get()
        for item in self.items: item.set_active_status(phase, "pending")
        self.refresh_table(); self.status_var.set(f"Reset all rows to pending in {phase} phase.")

    def delete_sent(self) -> None:
        phase = self.phase_var.get()
        before = len(self.items)
        self.items = [item for item in self.items if item.active_status(phase).lower() != "sent"]
        self.refresh_table(); self.status_var.set(f"Deleted {before - len(self.items)} sent rows in {phase} phase.")

    def clear_queue(self) -> None:
        if not self.items: return
        if not messagebox.askyesno(APP_TITLE, "Clear all queue rows? This does not delete your CSV file."): return
        self.items = []
        self.queue_path = None
        self.global_refs = {"product": [], "face": [], "outfit": [], "style": []}
        self.refresh_global_refs()
        self.refresh_table(); self.status_var.set("Queue and global refs cleared. No cached sent status remains.")

    def add_item(self) -> None:
        item = self.item_dialog()
        if item: self.items.append(item); self.refresh_table()

    def edit_item(self) -> None:
        idx = self.selected_index()
        if idx is None: return
        item = self.item_dialog(self.items[idx])
        if item: self.items[idx] = item; self.refresh_table()

    def remove_item(self) -> None:
        idx = self.selected_index()
        if idx is None: return
        self.items.pop(idx); self.refresh_table()

    def item_dialog(self, existing: QueueItem | None = None) -> QueueItem | None:
        dialog = tk.Toplevel(self); dialog.title("Project Queue Item"); dialog.geometry("880x680"); dialog.transient(self); dialog.grab_set()
        result: dict[str, QueueItem | None] = {"item": None}
        item = existing or QueueItem(profile_still=self.profile_var.get(), profile_motion="FLOW", still_delay_sec=self.delay_var.get())
        shot_var = tk.StringVar(value=item.shot_id); type_var = tk.StringVar(value=item.shot_type or "first_frame")
        still_profile_var = tk.StringVar(value=item.profile_still); motion_profile_var = tk.StringVar(value=item.profile_motion)
        still_delay_var = tk.IntVar(value=item.still_delay_sec); motion_delay_var = tk.IntVar(value=item.motion_delay_sec)
        still_output_var = tk.StringVar(value=item.still_output_name); motion_output_var = tk.StringVar(value=item.motion_output_name)
        first_frame_var = tk.StringVar(value=item.first_frame_path)
        attachments = list(item.attachments)
        frame = ttk.Frame(dialog, padding=12); frame.pack(fill="both", expand=True)
        for c in range(4): frame.columnconfigure(c, weight=1)
        ttk.Label(frame, text="Shot ID").grid(row=0, column=0, sticky="w"); ttk.Entry(frame, textvariable=shot_var).grid(row=0, column=1, sticky="ew", padx=6)
        ttk.Label(frame, text="Shot Type").grid(row=0, column=2, sticky="w"); ttk.Entry(frame, textvariable=type_var).grid(row=0, column=3, sticky="ew", padx=6)
        ttk.Label(frame, text="Still Profile").grid(row=1, column=0, sticky="w", pady=(6,0)); ttk.Entry(frame, textvariable=still_profile_var).grid(row=1, column=1, sticky="ew", padx=6, pady=(6,0))
        ttk.Label(frame, text="Motion Profile").grid(row=1, column=2, sticky="w", pady=(6,0)); ttk.Entry(frame, textvariable=motion_profile_var).grid(row=1, column=3, sticky="ew", padx=6, pady=(6,0))
        ttk.Label(frame, text="Still Prompt").grid(row=2, column=0, sticky="nw", pady=(10,0))
        still_text = tk.Text(frame, height=8, wrap="word"); still_text.grid(row=2, column=1, columnspan=3, sticky="nsew", padx=6, pady=(10,0)); still_text.insert("1.0", item.still_prompt)
        ttk.Label(frame, text="Motion Prompt").grid(row=3, column=0, sticky="nw", pady=(10,0))
        motion_text = tk.Text(frame, height=8, wrap="word"); motion_text.grid(row=3, column=1, columnspan=3, sticky="nsew", padx=6, pady=(10,0)); motion_text.insert("1.0", item.motion_prompt)
        frame.rowconfigure(2, weight=1); frame.rowconfigure(3, weight=1)
        ttk.Label(frame, text="First Frame Path").grid(row=4, column=0, sticky="w", pady=(8,0)); ttk.Entry(frame, textvariable=first_frame_var).grid(row=4, column=1, columnspan=2, sticky="ew", padx=6, pady=(8,0))
        def choose_first():
            f = filedialog.askopenfilename(title="Select first frame image")
            if f: first_frame_var.set(f)
        ttk.Button(frame, text="Choose", command=choose_first).grid(row=4, column=3, sticky="ew", padx=6, pady=(8,0))
        ttk.Label(frame, text="Attachments for Still").grid(row=5, column=0, sticky="nw", pady=(10,0))
        listbox = tk.Listbox(frame, height=5); listbox.grid(row=5, column=1, columnspan=3, sticky="nsew", padx=6, pady=(10,0))
        def fill_list():
            listbox.delete(0, "end")
            for p in attachments: listbox.insert("end", p)
        def add_files():
            files = filedialog.askopenfilenames(title="Select images/files")
            append_unique(attachments, files); fill_list()
        def remove_files():
            for row in reversed(list(listbox.curselection())): attachments.pop(row)
            fill_list()
        btns = ttk.Frame(frame); btns.grid(row=6, column=1, columnspan=3, sticky="w", padx=6, pady=(8,0))
        ttk.Button(btns, text="Add Files", command=add_files).pack(side="left", padx=(0,6)); ttk.Button(btns, text="Remove", command=remove_files).pack(side="left")
        ttk.Label(frame, text="Delay / Output").grid(row=7, column=0, sticky="w", pady=(10,0))
        ttk.Spinbox(frame, from_=1, to=1800, textvariable=still_delay_var, width=8).grid(row=7, column=1, sticky="w", padx=6, pady=(10,0))
        ttk.Spinbox(frame, from_=1, to=1800, textvariable=motion_delay_var, width=8).grid(row=7, column=2, sticky="w", padx=6, pady=(10,0))
        ttk.Entry(frame, textvariable=still_output_var).grid(row=8, column=1, sticky="ew", padx=6, pady=(6,0))
        ttk.Entry(frame, textvariable=motion_output_var).grid(row=8, column=2, sticky="ew", padx=6, pady=(6,0))
        ttk.Label(frame, text="Still output / Motion output").grid(row=8, column=0, sticky="w", pady=(6,0))
        footer = ttk.Frame(frame); footer.grid(row=9, column=0, columnspan=4, sticky="e", pady=(12,0))
        def save():
            if not shot_var.get().strip(): messagebox.showerror(APP_TITLE, "Shot ID is required."); return
            result["item"] = QueueItem(
                enabled=item.enabled, shot_order=item.shot_order, shot_id=shot_var.get().strip(), shot_type=type_var.get().strip(),
                category_th=item.category_th, scene_preset=item.scene_preset, scene_custom=item.scene_custom, style_preset=item.style_preset, aspect_ratio=item.aspect_ratio,
                profile_still=still_profile_var.get().strip(), profile_motion=motion_profile_var.get().strip(),
                still_prompt=still_text.get("1.0", "end").strip(), motion_prompt=motion_text.get("1.0", "end").strip(), attachments=list(attachments),
                product_ref_names=item.product_ref_names, face_ref_names=item.face_ref_names, outfit_ref_names=item.outfit_ref_names, style_ref_names=item.style_ref_names,
                still_delay_sec=int(still_delay_var.get()), motion_delay_sec=int(motion_delay_var.get()),
                still_output_name=still_output_var.get().strip(), first_frame_path=first_frame_var.get().strip(), motion_output_name=motion_output_var.get().strip(),
                still_status=item.still_status, motion_status=item.motion_status, notes=item.notes,
            )
            dialog.destroy()
        ttk.Button(footer, text="Cancel", command=dialog.destroy).pack(side="right", padx=(6,0)); ttk.Button(footer, text="Save", command=save).pack(side="right")
        fill_list(); dialog.wait_window(); return result["item"]

    def start(self) -> None:
        if self.busy: messagebox.showinfo(APP_TITLE, "Sender is already running."); return
        if not self.items: messagebox.showinfo(APP_TITLE, "Import or add queue items first."); return
        self.save_current_profile(); self.stop_event.clear(); self.busy = True
        self.worker = threading.Thread(target=self.worker_loop, daemon=True); self.worker.start(); self.status_var.set(f"Started {self.phase_var.get()} phase.")

    def stop(self) -> None:
        self.stop_event.set(); self.status_var.set("Stopping after current step...")

    def send_selected(self) -> None:
        idxs = self.selected_indices()
        if not idxs:
            idx = self.selected_index()
            if idx is None: return
            idxs = [idx]
        if self.busy: messagebox.showinfo(APP_TITLE, "Stop current run first."); return
        self.save_current_profile(); self.stop_event.clear(); self.busy = True
        self.worker = threading.Thread(target=lambda: self.worker_loop(idxs), daemon=True); self.worker.start()

    def worker_loop(self, only_indices: list[int] | None = None) -> None:
        phase = self.phase_var.get()
        try:
            indices = only_indices if only_indices is not None else list(range(len(self.items)))
            active_indices = [idx for idx in indices if self.items[idx].enabled.lower() not in {"no", "false", "0", "off"}]
            for pos, idx in enumerate(active_indices):
                if self.stop_event.is_set(): break
                item = self.items[idx]
                status = item.active_status(phase).lower()
                if self.skip_sent_var.get() and status == "sent":
                    self.ui_queue.put(("status", f"Skipped sent {phase}: {item.shot_id}")); continue
                if not item.active_prompt(phase).strip():
                    self.ui_queue.put(("status", f"Skipped empty {phase} prompt: {item.shot_id}")); continue
                self.ui_queue.put(("status", f"Sending {phase}: {item.shot_id}..."))
                item.set_active_status(phase, "sending"); self.ui_queue.put(("refresh", None)); self.autosave_queue()
                try:
                    self.send_item(item, phase)
                    item.set_active_status(phase, "sent"); self.ui_queue.put(("status", f"Sent {phase}: {item.shot_id}"))
                except StopRequested:
                    item.set_active_status(phase, "pending")
                    self.ui_queue.put(("status", STOP_MESSAGE))
                    self.write_log(f"{STOP_MESSAGE} during {phase} {item.shot_id}")
                    break
                except Exception as exc:
                    item.set_active_status(phase, "failed"); self.ui_queue.put(("status", f"Failed {phase} {item.shot_id}: {exc}")); self.write_log(f"Failed {phase} {item.shot_id}: {exc}")
                self.ui_queue.put(("refresh", None)); self.autosave_queue()
                if only_indices is None and pos != len(active_indices) - 1:
                    self.wait_or_stop(max(1, int(item.active_delay(phase))), f"Waiting after {item.shot_id}")
        except StopRequested:
            self.ui_queue.put(("status", STOP_MESSAGE))
            self.write_log(STOP_MESSAGE)
        finally:
            self.busy = False; self.ui_queue.put(("done", None))

    def send_item(self, item: QueueItem, phase: str) -> None:
        self.focus_countdown(item.shot_id)
        attachments = self.final_attachments(item, phase)
        self.ui_queue.put(("status", f"{item.shot_id} phase={phase} refs={len(attachments)}"))
        self.write_log(f"{item.shot_id} phase={phase} refs={len(attachments)} paths=" + " | ".join(attachments))
        if phase == "still" and not attachments:
            raise RuntimeError("Still phase has no reference image. Add global product/face refs or per-shot attachments first.")
        if phase == "motion" and not attachments:
            raise RuntimeError("Motion phase needs first_frame_path. Edit row and attach the generated still image first.")
        if self.click_each_var.get(): self.run_click_sequence(item.shot_id)
        if attachments:
            for file_index, ref_path in enumerate(attachments, start=1):
                if self.stop_event.is_set():
                    raise StopRequested(STOP_MESSAGE)
                retry(f"Set file clipboard {item.shot_id} ref {file_index}", lambda p=ref_path: set_file_clipboard_windows([p]))
                time.sleep(0.2); retry(f"Paste file {item.shot_id} ref {file_index}", send_ctrl_v)
                self.write_log(f"{item.shot_id}: pasted file {file_index}/{len(attachments)}: {ref_path}")
                if file_index < len(attachments): self.wait_or_stop(max(1, int(self.file_gap_var.get())), f"File gap for {item.shot_id} ref {file_index}")
            self.wait_or_stop(max(1, int(self.upload_var.get())), f"Upload wait for {item.shot_id}")
        prompt = item.active_prompt(phase).strip()
        retry(f"Set prompt clipboard {item.shot_id}", lambda: set_text_clipboard_windows(prompt))
        time.sleep(0.25); retry(f"Paste prompt {item.shot_id}", send_ctrl_v)
        time.sleep(0.3); retry(f"Press Enter {item.shot_id}", send_enter)
        self.write_log(f"Sent {phase} {item.shot_id}, chars={len(prompt)}, refs={len(attachments)}")

    def wait_or_stop(self, seconds: int, label: str) -> None:
        for remaining in range(seconds, 0, -1):
            if self.stop_event.is_set():
                raise StopRequested(STOP_MESSAGE)
            if remaining % 10 == 0 or remaining <= 5: self.ui_queue.put(("status", f"{label}: {remaining}s"))
            time.sleep(1)

    def log(self, msg: str) -> None:
        self.write_log(msg); self.status_var.set(msg)

    def write_log(self, msg: str) -> None:
        stamp = time.strftime("%Y-%m-%d %H:%M:%S")
        try: LOG_PATH.open("a", encoding="utf-8").write(f"[{stamp}] {msg}\n")
        except OSError: pass

    def drain_ui_queue(self) -> None:
        try:
            while True:
                kind, payload = self.ui_queue.get_nowait()
                if kind == "status": self.status_var.set(str(payload))
                elif kind == "refresh": self.refresh_table()
                elif kind == "done":
                    self.deiconify(); self.status_var.set("Queue done." if not self.stop_event.is_set() else "Stopped."); self.refresh_table()
        except queue.Empty:
            pass
        self.after(200, self.drain_ui_queue)


def main() -> None:
    app = PromptSenderApp(); app.mainloop()


if __name__ == "__main__":
    main()
